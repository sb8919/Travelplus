import os
import pandas as pd
from openpyxl import load_workbook

def merge_start():
    folder_path = '/root/web/Travel_plus_parsing_web/static/data/merge'
    # 업로드된 파일들을 데이터프레임으로 변환하여 리스트에 추가하기
    all_files = os.listdir(folder_path)
    excel_files = [f for f in all_files if f.endswith('.xlsx')]
    print(excel_files)
    dfs = []
    for file in excel_files:
        path = os.path.join(folder_path, file)
        wb = load_workbook(filename=path)
        sheet_names = wb.sheetnames
        for sheet_name in sheet_names:
            ws = wb[sheet_name]
            data = ws.values
            cols = next(data)
            df = pd.DataFrame(data, columns=cols)
            dfs.append(df)

    # 모든 DataFrame을 합치기
    result = pd.concat(dfs)

    result = result.drop_duplicates(subset=['해시태그'])
    result = result.drop_duplicates(subset=['본문'])
    result = result[result['해시태그'] != '[]']
    result = result.dropna(subset=['본문', '해시태그'],how='all')
    # 결과 저장
    result.to_excel(os.path.join(folder_path+'/complete', 'complete.xlsx'), index=False)
    return print('병합 완료!')
