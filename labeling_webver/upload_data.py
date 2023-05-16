import pandas as pd
import dbcl
import xlrd
from openpyxl import load_workbook
import re

labeling = dbcl.DBconnector('label_Data')

def insert_data_to_db():
    wb = load_workbook(filename = './static/data/upload/upload.xlsx')
    sheet = wb['Sheet1']
    
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row = list(row)
        if row[3]==None:
            row[3] = ''
        if row[2]==None:
            row[2] = ''
        row[4] = row[4].replace("'",'').replace('[','').replace(']','')
        row[3] = row[3].replace('(',' ').replace(')',' ').replace("‘",'').replace('’','').replace("🪨",'').replace("",'')
        row[0], row[1], row[2], row[3], row[4] = row[3], row[4], row[1], 'None', row[2]
        try:
            sql = f"INSERT INTO labeling(Contents, HashTag, img_url, label, location) VALUES {tuple(row)};"
            labeling.insert(sql)
        except:
            pass
    
    